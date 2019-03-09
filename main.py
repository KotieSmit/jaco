from openpyxl import load_workbook, cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import Workbook
from subprocess import Popen, PIPE
from time import strptime, sleep
from datetime import datetime
import requests
import os
import json

from configparser import ConfigParser
 
 #!/usr/bin/python
import psycopg2
from config import config
 
def connect():
    """ Connect to the PostgreSQL database server """
    conn = None
    try:
        # read connection parameters
        params = config()
 
        # connect to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        conn = psycopg2.connect(**params)
 
        # create a cursor
        cur = conn.cursor()
        
        # execute a statement
        print('PostgreSQL database version:')
        cur.execute('SELECT version()')
 
        # display the PostgreSQL database server version
        db_version = cur.fetchone()
        print(db_version)
       
     # close the communication with the PostgreSQL
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print('Database connection closed.')
            
            
def insert_sites(sites):
    # sample: sites = [(3, "klerksdorp", "10.0.0.1")]
    sql = "INSERT INTO public.sites(id, name, ip) VALUES(%s, %s, %s)"
    print(sql)
    conn = None
    try:
        # read database configuration
        params = config()
        # connect to the PostgreSQL database
        conn = psycopg2.connect(**params)
        # create a new cursor
        cur = conn.cursor()
        # execute the INSERT statement
        cur.executemany(sql,sites)
        # commit the changes to the database
        conn.commit()
        # close communication with the database
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()

def get_site_id(ip):
    """ query data from the vendors table """
    conn = None
    try:
        params = config()
        conn = psycopg2.connect(**params)
        cur = conn.cursor()
        cur.execute("SELECT *  FROM sites WHERE ip = '" + ip + "'")
        row = cur.fetchone()
 
        while row is not None:
            return row[0]
 
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()


def insert_stats(stats):

    sql = 'INSERT INTO public.stats("site_id", date, tx, rx) VALUES(%s, %s, %s, %s)'
    conn = None
    try:
        # read database configuration
        params = config()
        # connect to the PostgreSQL database
        conn = psycopg2.connect(**params)
        # create a new cursor
        cur = conn.cursor()
        # execute the INSERT statement
        cur.executemany(sql,stats)
        # commit the changes to the database
        conn.commit()
        # close communication with the database
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()

def readIndex():
    wb = Workbook()
    wb = load_workbook("Btest.xlsx", keep_vba=True)
    ws = wb['Index']

    for row in ws.rows:
        if row[0].value != None:

            site_id  = row[0].value
            ip = row[1].value
            name = row[2].value

            print(type(site_id))
            print(ip)
            print(name)
            # print(link)
            insert_sites([(site_id, name, ip)])


def findSiteById():
    wb = Workbook()
    wb = load_workbook("Btest.xlsx", keep_vba=True)
    ws = wb['Index']
    site =None
    for row in ws.rows:
        if row[0].value != None:
            name = row[2].hyperlink.location.split("!")[0].strip()
            if "'" in name:
                name = name.split("'")[1]
            # print("-" + name + "-")
            try:
                # First try as name appears in hyperlink
                site = wb[name]
            except Exception as identifier:
                pass
                try:
                # Then strip out spaces at start and end of name
                    site = wb[name].strip()
                except Exception as identifier:
                    pass
                    try:
                        #if all else fails try by name in the sheet
                        name = row[2].value
                        site = wb[name]
                    except Exception as identifier:
                        print(name + " sheet not found")
                    
        site_id  = row[0].value

        if site != None and site_id != None:
            for site_row in site.rows:
                if site_row[2].value != None and site_row[2].value != "TX":
                    date = site_row[1].value
                    tx = site_row[2].value
                    rx = site_row[3].value
                    print(site_id)
                    # print(date)
                    # print(tx)
                    # print(rx) 
                
        
                    insert_stats([(site_id, date, tx, rx)])

                    


            

    # for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row)):
    #         for col in ws.iter_cols(min_col=1,max_col=1):
    #                 for cell in row:
    #                     cv = cell.value
                        
    #                     # print (cell.row,cell.column)
    #                     # print (cell.row,cell.column)
    #                     if self.data["ip"] == cv:
    #                             # print (cell.value)
    #                             # print (cell.coordinate)
    #                             self.siteName = cell.hyperlink.location.split("'")[1]
    #                             # return index[get_column_letter( c.col_idx + 1) + str(c.row)].value
    #                             break
    #                     else:
    #                             # print ("Not found")
    #                             pass

def insertFromfile():    
    for fname in os.listdir("import"):
        if fname.endswith(".txt"): 
            with open("import/" + fname) as f:
                content = f.readlines()
                content = [x.strip() for x in content]    
            for line in content:
                if "rx-total-average" in line and "10.240.0.65" not in line:
                    try:
                        #results = line.split("tx-total-average: ")
                        tx =  int(float(line.split("tx-total-average: ")[1].split(" Mbps")[0]))
                        rx =  int(float(line.split("rx-total-average: ")[1].split(" Mbps")[0]))
                        date = strptime(line.split(" ")[0],'%b/%d/%Y')
                        date = str(date.tm_year) + "-" + str(date.tm_mon) + "-" + str(date.tm_mday)
                        ip = line.split(" ")[2]
                        
                        site_id = get_site_id(ip)
                        sql = 'INSERT INTO public.stats("site_id", date, tx, rx) VALUES(%s, %s, %s, %s)'
                        insert_stats([(site_id, date, tx, rx)])
                    except expression as err:
                        print(err)
                        print("FAILED to insert follwing data:")
                        print('Date: {2} Ip: {3}  Tx: {0} Rx: {1}'.format(tx, rx, date, ip))
                        pass
            print("Imported file: " + fname)
            # Move imprted data to a imported folder
            try:
                today = datetime.now()
                today= today.strftime('%Y%m%d')
                # Check if folder exist
                if not os.path.isdir("import/imported/" + today):
                    #Create folder
                    os.mkdir("import/imported/" + today)
                    print("Created folder: " + today)
                os.rename("import/" + fname, "import/imported/" + today + "/"+ fname)
                print("Moved '" + fname + "' to imported folder.")
            except Exception as err:
                print(err)
                pass



if __name__ == '__main__':
    print("Started import server.")
    # readIndex()
    #findSiteById()
  
  
    while True:
        insertFromfile()
        sleep(5)
    # get_site_id('10.240.63.250')