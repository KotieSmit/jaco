from openpyxl import load_workbook, cell
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl import Workbook
from subprocess import Popen, PIPE
import requests
import os
import json

class TWK:
    fileDir = ""
    file = ""
    data = {"date":"2019-02-13", "ip": "10.240.63.253", "tx": "777.867", "rx": "999.867"}
    

    wb= None
    index = None
    siteSheet = None
    siteName = None
    txRxInputRow = None
    txCol = "C"
    rxCol = "D"

    def __init__(self):
        print("start")
        self.openFile()
        self.updateSheet(self.data)
        self.wb.save("bla.xlsm")
        print("end")

    def openFile(self):
        wb = Workbook()
        self.wb = load_workbook("Btest.xlsx", keep_vba=True)
        self.index = self.wb['Index']

    def findSiteByIp(self):
        index = self.index
        for row in index.iter_rows('A{}:A{}'.format(index.min_row,index.max_row)):
               for col in index.iter_cols(min_col=1,max_col=1):
                       for cell in row:
                            cv = cell.value
                            
                            # print (cell.row,cell.column)
                            # print (cell.row,cell.column)
                            if self.data["ip"] == cv:
                                    # print (cell.value)
                                    # print (cell.coordinate)
                                    self.siteName = cell.hyperlink.location.split("'")[1]
                                    # return index[get_column_letter( c.col_idx + 1) + str(c.row)].value
                                    break
                            else:
                                    # print ("Not found")
                                    pass


    def findDataEntryPoint(self):
        sheet = self.siteSheet
        for row in sheet.iter_rows('B{}:B{}'.format(sheet.min_row,sheet.max_row)):
               for col in sheet.iter_cols(min_col=1,max_col=1):
                       for cell in row:
                            cv = str(cell.value)
                            if cv != None:
                                # print cv.split(" ")[0]
                                if cv.split(" ")[0]== self.data["date"] :
                                    print("findDataEntryPoint")
                                    print (cell.coordinate)
                                    print (cell.value)
                                    # self.siteName = cell.hyperlink.location.split("'")[1]
                                    self.txRxInputRow = cell.row
                                    break
                                
        if self.txRxInputRow == None:
             print ("Not found - " + self.data["date"])
                                        


    def getSiteSheet(self):
        self.siteSheet = self.wb[self.siteName]


    def updateSheet(self, data):
        print data["rx"]
        
        self.findSiteByIp()
        self.getSiteSheet()
        self.findDataEntryPoint()
        sheet = self.siteSheet

        sheet[self.txCol + str(self.txRxInputRow)].value = data["tx"]
        sheet[self.rxCol + str(self.txRxInputRow)].value = data["rx"]

twk = TWK()